# Kieran Hulsman: March-May 2020
if __name__ == '__main__': 
    from datetime import datetime
    from openpyxl import load_workbook
    import re
    import smtplib
    import ssl
    import sys
    import secrets


    wb = load_workbook("userDatabase.xlsx")
    sheet1 = wb.active

    datetime_raw = str(datetime.now())


    class User:
        def __init__(self, first_name, last_name, email, password):
            self.first_name = first_name
            self.last_name = last_name
            self.email = email
            self.password = password


    def yes_no_question(question):
        answer_is_valid = False
        while answer_is_valid is False:
            prompt = input(question).lower()
            valid_answers = ("y", "yes", "n", "no")
            if prompt in valid_answers:
                if prompt in valid_answers[0:2]:
                    return True
                else:
                    return False
            else:
                answer_is_valid = False
                print("---\nPlease enter \"y\" or \"n\"")


    def input_required(prompt_text):
        input_just_enter = True
        while input_just_enter:
            prompt = input(prompt_text)
            if prompt == "":
                print("---\nField required")
            else:
                input_just_enter = False
        return prompt


    def name_reformat(unformatted_name):
        formatted_name = unformatted_name[0].upper() + unformatted_name[1:].lower()
        return formatted_name


    def name_splitter(prompt):
        while IndexError:
            contains_digits_outer = True
            while contains_digits_outer:
                unformatted_name_prompt = input_required(prompt)
                contains_digits_loop = False
                for character in unformatted_name_prompt:
                    if character.isdigit():
                        contains_digits_loop = True
                    else:
                        pass
                if contains_digits_loop:
                    print("---\nNot a real name")
                else:
                    contains_digits_outer = False
            try:
                name = [name_reformat(unformatted_name_prompt.split()[0]),
                        name_reformat(unformatted_name_prompt.split()[1])]
                break
            except IndexError:
                print("---\nPlease enter your first AND last name (make sure there's a space in between them)")
        return name


    def valid_email_address(prompt):
        valid_email = False
        while not valid_email:
            email = input_required(prompt)

            regex = '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'
            if re.search(regex, email):
                valid_email = True
            else:
                print("---\nInvalid email address")
        return email


    def cell_value_finder(attribute):
        if attribute.lower() == "first name":
            col = "A"
        elif attribute.lower() == "last name":
            col = "B"
        elif attribute.lower() == "email":
            col = "C"
        elif attribute.lower() == "password":
            col = "D"
        else:
            return "INVALID PARAMETER - not an attribute"

        cell_value = sheet1["{}{}".format(col, str(row_number))].value
        return cell_value


    def formatted_datetime():

        def formatted_date():
            month_raw = datetime_raw[5:7]
            day_raw = datetime_raw[8:10]
            year_raw = datetime_raw[0:4]
            
            month = {
                "01": "January",
                "02": "February",
                "03": "March",
                "04": "April",
                "05": "May",
                "06": "June",
                "07": "July",
                "08": "August",
                "09": "September",
                "10": "October",
                "11": "November",
                "12": "December",
                }

            def day_suffix(day):
                if int(day) in range(4, 21) or int(day) > 23:
                    return "th"
                else:
                    if day[1] == "1":
                        return "st"
                    elif day[1] == "2":
                        return "nd"
                    else:
                        return "rd"

            return "{} {}{} {}".format(
                month.get(month_raw), 
                int(day_raw), 
                day_suffix(day_raw), 
                year_raw
                )


        def formatted_time():
            hour24_str = datetime_raw[11:13]
            hour24_int = int(hour24_str)
            minutes = datetime_raw[13:16]

            if hour24_int > 12:
                hour12 = hour24_int - 12
                return "{}{}PM".format(hour12, minutes)
            else: 
                if hour24_str == "00":
                    return "12{}AM".format(minutes)
                else:
                    return "{}{}AM".format(hour24_int, minutes)


        return "{}, at {}".format(formatted_date(), formatted_time())


    def account_info(greeting, first_name, last_name, email, password, is_email_body):
        if is_email_body:
            return ("{} {},"
                    "\n\nHere's your profile info:"
                    "\n\tNAME: {} {}"
                    "\n\tEMAIL: {}"
                    "\n\tPASSWORD: {}").format(
                        name_reformat(greeting), 
                        first_name, 
                        first_name, 
                        last_name, 
                        email, 
                        password
                        )
        else:
            print("\n---\n\n{} {}!"
                "\n\nHere's your profile info:"
                "\n\tNAME: {} {}"
                "\n\tEMAIL: {}"
                "\n\tPASSWORD: {}".format(
                        name_reformat(greeting), 
                        first_name, 
                        first_name, 
                        last_name, 
                        email, 
                        password
                        )
                        )


    def send_email(user_email, body, header):
        email_authentication_errors_file = open("emailAuthenticationErrors.txt", "w")
        try: 
            port = 465
            smtp_server = "smtp.gmail.com"
            sender_email = secrets.sender_email
            sender_password = secrets.sender_password
            message = "Subject: {}\n\n{}".format(body, header)

            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
                server.login(sender_email, sender_password)
                server.sendmail(sender_email, user_email, message)
            email_authentication_status = "No authentication error"
        except smtplib.SMTPAuthenticationError:
            email_authentication_status = "smtplib.SMTPAuthenticationError"
        email_authentication_errors_file.writelines("{}\n{}".format(email_authentication_status, datetime_raw))
        email_authentication_errors_file.close()


    email_column = sheet1["C"]
    user_emails = []
    for cell in email_column:
        user_emails.append(cell.value)

    has_account = yes_no_question("Do you have an account? (y/n): ")
    if has_account:
        wants_login = yes_no_question("Do you want to login? (y/n): ")
        if wants_login:
            valid_email_login = False
            while not valid_email_login:
                email_login = valid_email_address("Enter email: ")
                reformatted_email_login = email_login.lower()
                if reformatted_email_login in user_emails:
                    row_number = user_emails.index(reformatted_email_login) + 1
                    valid_password_login = False
                    out_of_password_attempts = False
                    password_attempt_count = 0
                    password_attempt_limit = 3
                    while not valid_password_login and not out_of_password_attempts:
                        if password_attempt_count < password_attempt_limit:
                            password_login = input_required("Enter password: ")
                            password_attempt_count += 1
                        else:
                            print("Out of login attempts, please try again later")
                            password_attempt_header = "Potential security threat on your account"
                            password_attempt_body = "Dear {}, " \
                                                    "\n\nA login attempt was just made with your email address." \
                                                    "\n\nWhen in happened:" \
                                                    "\n{}".format(
                                                        cell_value_finder("first name"), formatted_datetime()
                                                        )
                            send_email(reformatted_email_login, password_attempt_header, password_attempt_body)
                            sys.exit()
                        if password_login == sheet1["D{}".format(str(row_number))].value:
                            account_info("hello",
                                        cell_value_finder("first name"),
                                        cell_value_finder("last name"),
                                        cell_value_finder("email"),
                                        cell_value_finder("password"),
                                        False)
                            valid_password_login = True
                            wants_edit_profile = yes_no_question("\nEdit profile? (y/n): ")
                            if wants_edit_profile:
                                edit_name_text = "Change name? ({} {}): ".format(
                                    cell_value_finder("first name"), 
                                    cell_value_finder("last name")
                                    )
                                wants_edit_name = yes_no_question(edit_name_text)
                                old_name = "{} {}".format(
                                    cell_value_finder("first name"),
                                    cell_value_finder("last name")
                                    )
                                if wants_edit_name:
                                    edited_full_name_prompt = name_splitter("Enter new name: ")
                                    edited_first_name = edited_full_name_prompt[0]
                                    edited_last_name = edited_full_name_prompt[1]
                                    if "{} {}".format(edited_first_name, edited_last_name) == old_name:
                                        print("---\nNew name is the the same as the old one")
                                    else:
                                        pass
                                    edited_full_name_value = [edited_first_name, edited_last_name]
                                else:
                                    edited_full_name_value = [cell_value_finder("first name"),
                                                            cell_value_finder("last name")]
                                edit_email_text = "Change email? ({}): ".format(cell_value_finder("email"))
                                wants_edit_email = yes_no_question(edit_email_text)
                                old_email = cell_value_finder("email")
                                if wants_edit_email:
                                    edited_email = valid_email_address("Enter new email: ")
                                    if edited_email == old_email:
                                        print("---\nNew email is the same as the old one")
                                    else:
                                        pass
                                    sheet1.cell(row_number, 3).value = edited_email.lower()
                                else:
                                    edited_email = old_email
                                wants_edit_password = \
                                    yes_no_question("Change password? ({}): ".format(cell_value_finder("password")))
                                old_password = cell_value_finder("password")
                                if wants_edit_password:
                                    edited_password = input_required("Enter new password: ")
                                    if edited_password == old_password:
                                        print("---\nNew password is the same as the old one")
                                    else:
                                        pass
                                    sheet1.cell(row_number, 4).value = edited_password
                                else:
                                    edited_password = old_password
                                edited_user_object = User(edited_full_name_value[0],
                                                        edited_full_name_value[1],
                                                        edited_email,
                                                        edited_password)
                                edited_user_attributes = [edited_user_object.first_name,
                                                        edited_user_object.last_name,
                                                        edited_user_object.email,
                                                        edited_user_object.password]
                                print("\nProfile information successfully updated")
                                for attribute in range(4):
                                    sheet1.cell(row_number, attribute + 1).value = edited_user_attributes[attribute]
                                edit_header = "Updated Profile Information"
                                edit_body = account_info("hello",
                                                        edited_user_attributes[0],
                                                        edited_user_attributes[1],
                                                        edited_user_attributes[2],
                                                        edited_user_attributes[3],
                                                        True)
                                send_email(edited_email, edit_header, edit_body)
                            else:
                                print("\nOkay then. Have a great day!")
                        else:
                            if not out_of_password_attempts:
                                print("---\nPassword doesn't match email address")
                            else:
                                pass
                        valid_email_login = True
                else:
                    print("---\nThere are no accounts registered with that email address")
        else:
            print("\nOkay then. Have a great day!")
    else:
        wants_account = yes_no_question("Want to create an account? (y/n): ")
        if wants_account:
            full_name_prompt = name_splitter("Enter your name: ")
            first_name_prompt = full_name_prompt[0]
            last_name_prompt = full_name_prompt[1]
            new_email = False
            while not new_email:
                email_prompt = valid_email_address("Enter your email: ")
                if email_prompt not in user_emails:
                    new_email = True
                else:
                    print("---\nEmail address already registered with an account")
            correct_password_reentry = False
            while correct_password_reentry is False:
                password_prompt = input_required("Create a password: ")
                password_verification = input_required("Re-enter password: ")
                if password_prompt == password_verification:
                    correct_password_reentry = True
                else:
                    print("---\nHmm... those passwords don't match")
            reformatted_email = email_prompt.lower()
            account_info("welcome",
                        first_name_prompt,
                        last_name_prompt,
                        reformatted_email,
                        password_prompt,
                        False)
            new_user_object = User(first_name_prompt, last_name_prompt, reformatted_email, password_prompt)
            new_user_attributes = [
                new_user_object.first_name,
                new_user_object.last_name,
                new_user_object.email,
                new_user_object.password
                ]
            sheet1.append(new_user_attributes)
            welcome_header = "Thanks for signing up!"
            welcome_body = account_info(
                "welcome",
                new_user_attributes[0],
                new_user_attributes[1],
                new_user_attributes[2],
                new_user_attributes[3],
                True
                )
            send_email(reformatted_email, welcome_header, welcome_body)
        else:
            print("\nOkay then. Have a great day!")
    wb.save("userDatabase.xlsx")
else:
    print("The file being fun isn't the original file. Because of this, the file won't run properly.")
